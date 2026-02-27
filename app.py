# app.py
import io
import random
import zipfile
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Squirrel Bingo Generator", page_icon="🐿️")

st.title("Squirrel Bingo Generator 🐿️")
st.caption("Edit the list, then generate beautifully chaotic cards.")

DEFAULT_ITEMS = """Take a nap.
Find some nature.
Find your keys (or wallet or phone)
Drink water.
Listen to music.
Talk to a friendly voice.
Don’t buy that shiny thing.
Reframe something bad that happened into good.
Snuggle something (stuffy, pet, human)
Have you eaten?
BIG stretch!
Take a shower (or wash your face).
Stop. Take 3 deep breaths.
Have a healthy snack.
Work out.
Do something creative.
Close a social media app.
Clean or declutter something.
Meditate.
Arrive early to an appointment.
Drink some water.
Write about something that happened.
Take your meds OR body double.
Check tomorrow’s (or next week’s) schedule.
"""

if "items_text" not in st.session_state:
    st.session_state.items_text = DEFAULT_ITEMS

if st.button("Reset to default deck"):
    st.session_state.items_text = DEFAULT_ITEMS

raw = st.text_area(
    "Self-care options (one per line). Need at least 24 unique items.",
    height=320,
    key="items_text",
)

free = st.text_input("Free center square", "🐿️")
n_cards = st.number_input("How many cards?", min_value=1, max_value=50, value=7)

# --- Parse & clean input ---
lines = [line.strip() for line in raw.splitlines()]
items = [x for x in lines if x]  # remove empty

# de-dupe (preserve order)
items = list(dict.fromkeys(items))

def make_card(all_items):
    """Return a strict 5x5 list-of-lists with center replaced by `free`."""
    if len(all_items) < 24:
        raise ValueError("Need at least 24 unique items to generate a 5x5 card.")

    # Choose 24 unique items for the non-center squares
    chosen = random.sample(all_items, 24)
    random.shuffle(chosen)

    # Fill 5x5 with a center hole at (2,2)
    grid = []
    k = 0
    for r in range(5):
        row = []
        for c in range(5):
            if r == 2 and c == 2:
                row.append(free)
            else:
                row.append(chosen[k])
                k += 1
        grid.append(row)

    # sanity check
    assert len(grid) == 5 and all(len(r) == 5 for r in grid)
    return grid

def card_to_xlsx_bytes(card, title):
    wb = Workbook()
    ws = wb.active
    ws.title = title

    for r in range(5):
        for c in range(5):
            cell = ws.cell(row=r + 1, column=c + 1, value=card[r][c])
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 28
    for row in range(1, 6):
        ws.row_dimensions[row].height = 60

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

if st.button("Generate"):
    if len(items) < 24:
        st.error(f"You have {len(items)} unique items; need at least 24.")
        st.stop()

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as z:
        for i in range(1, int(n_cards) + 1):
            card = make_card(items)
            xlsx = card_to_xlsx_bytes(card, f"Card {i}")
            z.writestr(f"Squirrel_Bingo_Card_{i}.xlsx", xlsx)

    st.download_button(
        "Download ZIP of Excel cards",
        data=zip_buf.getvalue(),
        file_name="squirrel_bingo_cards.zip",
        mime="application/zip",
    )
